#!/usr/bin/env python3
import openpyxl
import psycopg2
import os
from collections import defaultdict

# Connessione al database
conn = psycopg2.connect(
    host="localhost",
    database="assessment_db",
    user="assessment_user",
    password=os.getenv("DB_PASSWORD", "your_password")
)
cur = conn.cursor()

# Leggi Excel
wb = openpyxl.load_workbook('file di controllo.xlsx')
ws = wb.active

# Mappa categorie -> colonne
category_cols = {
    'Governance': (2, 6),  # B-F (5 dimensioni)
    'Monitoring & Control': (8, 11),  # H-K (4 dimensioni)
    'Technology': (13, 15),  # M-O (3 dimensioni)
    'Organization': (17, 18)  # Q-R (2 dimensioni)
}

# Leggi headers (riga 2)
headers = {}
for cat, (start, end) in category_cols.items():
    headers[cat] = [ws.cell(2, col).value for col in range(start, end + 1) if ws.cell(2, col).value and 'note' not in str(ws.cell(2, col).value).lower()]

print("=== HEADERS ===")
for cat, dims in headers.items():
    print(f"{cat}: {len(dims)} dimensioni")
    for d in dims:
        print(f"  - {d}")

# Leggi dati Excel
expected_data = []
current_process = None

for row_idx in range(3, ws.max_row + 1):
    activity_name = ws.cell(row_idx, 1).value
    
    if not activity_name:
        continue
    
    # Check se è un nome di processo (tutte le altre celle vuote)
    is_process = all(ws.cell(row_idx, col).value is None for col in range(2, 20))
    
    if is_process:
        current_process = activity_name
        print(f"\n--- PROCESSO: {current_process} ---")
        continue
    
    if not current_process:
        continue
    
    # Leggi scores per ogni categoria
    for category, (start_col, end_col) in category_cols.items():
        for dim_idx, col in enumerate(range(start_col, end_col + 1)):
            header = ws.cell(2, col).value
            if not header or 'note' in str(header).lower():
                continue
            
            score = ws.cell(row_idx, col).value
            if score is not None:
                expected_data.append({
                    'process': current_process,
                    'activity': activity_name,
                    'category': category,
                    'dimension': header,
                    'score': score
                })

print(f"\n=== DATI EXCEL ===")
print(f"Totale righe attese: {len(expected_data)}")

# Query database (usa il session_id corretto)
print("\n=== CONFRONTO CON DATABASE ===")
print("Quale session_id vuoi verificare?")

cur.close()
conn.close()
wb.close()
